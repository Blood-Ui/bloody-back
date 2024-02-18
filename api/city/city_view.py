from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from django.db import transaction
from tempfile import NamedTemporaryFile
from io import BytesIO
from django.http import FileResponse
from openpyxl import Workbook
from openpyxl.styles import Font

from api.models import City, District
from api.utils import CustomResponse, get_user_id, RoleList, allowed_roles, get_excel_data
from .city_serializer import CityCreateSerializer, CityListSerializer, CityUpdateSerializer, CityDropDownSerializer

class CityAPIView(APIView):
    permission_classes = [IsAuthenticated]

    @allowed_roles([RoleList.ADMIN.value])
    def post(self, request):
        user_id = get_user_id(request)
        serializer = CityCreateSerializer(data=request.data, context={'request': request, 'user_id': user_id})
        if serializer.is_valid():
            serializer.save()
            return CustomResponse(message="successfully created city", data=serializer.data).success_response()
        return CustomResponse(message="failed to create city", data=serializer.errors).failure_reponse()
    
    @allowed_roles([RoleList.ADMIN.value])
    def get(self, request):
        cities = City.objects.all()
        serializer = CityListSerializer(cities, many=True)
        return CustomResponse(message="successfully obtained cities", data=serializer.data).success_response()
    
    @allowed_roles([RoleList.ADMIN.value])
    def patch(self, request, city_id):
        user_id = get_user_id(request)
        if not City.objects.filter(id=city_id).exists():
            return CustomResponse(message="city does not exist").failure_reponse()
        city = City.objects.get(id=city_id)
        serializer = CityUpdateSerializer(city, data=request.data, context={'request': request, 'user_id': user_id})
        if serializer.is_valid():
            serializer.save()
            return CustomResponse(message="successfully updated city", data=serializer.data).success_response()
        return CustomResponse(message="failed to update city", data=serializer.errors).failure_reponse()
    
    @allowed_roles([RoleList.ADMIN.value])
    def delete(self, request, city_id):
        if not City.objects.filter(id=city_id).exists():
            return CustomResponse(message="city does not exist").failure_reponse()
        city = City.objects.get(id=city_id)
        city.delete()
        return CustomResponse(message="successfully deleted city").success_response()
    
class CityDropDownView(APIView):
    permission_classes = [IsAuthenticated]

    @allowed_roles([RoleList.ADMIN.value])
    def get(self, request, district_id):
        if not District.objects.filter(id=district_id).exists():
            return CustomResponse(message="district does not exist").failure_reponse()
        if not City.objects.filter(district=district_id).exists():
            return CustomResponse(message="No city is present for this district").failure_reponse()
        cities = City.objects.filter(district=district_id)
        serializer = CityDropDownSerializer(cities, many=True)
        return CustomResponse(message="successfully obtained cities", data=serializer.data).success_response()
    
class CityBulkImportAPIView(APIView):
    permission_classes = [IsAuthenticated]

    @allowed_roles([RoleList.ADMIN.value])
    def post(self, request):
        try:
            excel_file = request.FILES["cities"]
        except:
            return CustomResponse(message="file not found").failure_reponse()
        if not excel_file.name.endswith('.xlsx'):
            return CustomResponse(message="file type not supported").failure_reponse()
        excel_data = get_excel_data(excel_file)
        
        headers = ['name', 'district']
        if not excel_data:
            return CustomResponse(message="The file is empty.").failure_reponse()
        for header in headers:
            if header not in excel_data[0]:
                return CustomResponse(message=f"Please provide the {header} in the file.").failure_reponse()
            
        for index, data in enumerate(excel_data[1:]):
            district_name = data.pop('district')
            if not District.objects.filter(name=district_name).exists():
                return CustomResponse(message=f"In row index {index + 2} given district does not exist").failure_reponse()
            district = District.objects.get(name=district_name)
            data['district'] = district.id
        user_id = get_user_id(request)
        serializer = CityCreateSerializer(data=excel_data[1:], context={'request': request, 'user_id': user_id}, many=True)
        with transaction.atomic():
            if serializer.is_valid():
                if len(serializer.data) != len(excel_data[1:]):
                    transaction.set_rollback(True)
                    return CustomResponse(message="something went wrong, please try again", data=serializer.errors).failure_reponse()
                serializer.save()
                return CustomResponse(message="successfully created cities", data=serializer.data).success_response()
        errors_with_indices = []
        for index, error in enumerate(serializer.errors):
            errors_with_indices.append({"row_index": index + 2, "error": error if error else "no error"})  # Adjust index for headers
        return CustomResponse(message="failed to import cities", data=errors_with_indices).failure_reponse()
    
class CityBaseTemplateAPIView(APIView):
    permission_classes = [IsAuthenticated]

    @allowed_roles([RoleList.ADMIN.value])
    def get(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(['name', 'district'])
        # Set column headers font as bold
        bold_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = bold_font
        # Set column width
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 35
        ws = wb.create_sheet('Data Definitions')
        ws.append(['district'])
        # set column header as bold and set width
        for cell in ws[1]:
            cell.font = bold_font
        ws.column_dimensions['A'].width = 35
        districts = District.objects.all().values_list('name', flat=True)
        data = {'district': districts}
        # Write data column-wise
        for col_num, (col_name, col_values) in enumerate(data.items(), start=1):
            for row, value in enumerate(col_values, start=2):
                ws.cell(row=row, column=col_num, value=value)

        with NamedTemporaryFile() as tmp:
            tmp.close()  # with statement opened tmp, close it so wb.save can open it
            wb.save(tmp.name)
            with open(tmp.name, 'rb') as f:
                f.seek(0)
                new_file_object = f.read()
        return FileResponse(BytesIO(new_file_object), as_attachment=True, filename='city_base_template.xlsx')