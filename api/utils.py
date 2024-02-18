from rest_framework.response import Response
from rest_framework import status
from enum import Enum
from rest_framework_simplejwt.authentication import JWTAuthentication
import io
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from django.http import FileResponse
from tempfile import NamedTemporaryFile
from io import BytesIO


class CustomResponse():
    def __init__(self, message=None, data=None):
        self.message = {} if message is None else message
        self.data = {} if data is None else data

    def success_response(self):
        self.error = False
        self.status_code = status.HTTP_200_OK
        return Response({"error": self.error, "message": self.message, "data": self.data, "status_code": self.status_code}, status=self.status_code)
    
    def failure_reponse(self):
        self.error = True
        self.status_code = status.HTTP_400_BAD_REQUEST
        return Response({"error": self.error, "message": self.message, "data": self.data, "status_code": self.status_code}, status=self.status_code)
    
def get_user_id(request):
    JWT_authenticator = JWTAuthentication()
    response = JWT_authenticator.authenticate(request)
    if response is not None:
        # unpacking
        user , token = response
        user_id = token.payload['user_id']
        return user_id
    return None

def get_user_role(request):
    JWT_authenticator = JWTAuthentication()
    response = JWT_authenticator.authenticate(request)
    if response is not None:
        # unpacking
        user , token = response
        user_role = token.payload['roles']
        return user_role
    return None

class RequestStatus(Enum):
    OPEN = 'open'
    INPROGRESS = 'in_progress'
    APPROVED = 'approved'
    CANCELLED = 'cancelled'
    CLOSED = 'closed'

    @classmethod
    def get_all_values(cls):
        return [member.value for member in cls]
    
class RoleList(Enum):
    ADMIN = 'admin'
    INCHARGE = 'incharge'
    NORMAL_USER = 'normal_user'

    @classmethod
    def get_all_values(cls):
        return [member.value for member in cls]
    
def allowed_roles(allowed_roles):
    def decorator(func):
        def wrapper(obj, request, *args, **kwargs):
            user_roles = get_user_role(request) 
            flag = False
            for user_role in user_roles:
                if user_role in allowed_roles:
                    return func(obj, request, *args, **kwargs)
            if not flag:
                return CustomResponse(message="You don't have permission to perform this action").failure_reponse()
        return wrapper
    return decorator

def get_excel_data(excel_file):
    # Read the file in a memory-efficient way
    file_buffer = io.BytesIO(excel_file.read())
    wb = load_workbook(file_buffer)
    worksheet = wb.active
    excel_data = []
    for row in worksheet.iter_rows(values_only=True):
        # Check for empty rows before processing
        if any(cell for cell in row):  # Use zip_longest to handle potentially different header/data lengths
            row_data = {header.value: cell for header, cell in zip(worksheet[1], row) if cell is not None}
            excel_data.append(row_data)

    return excel_data

def generate_excel_template(sheet_names, headers, data_dict, column_widths, filename):
    wb = Workbook()
    bold_font = Font(bold=True)
    for i, sheet_name in enumerate(sheet_names):
        ws = wb.create_sheet(sheet_name)
        # Write headers
        ws.append(headers[i])
        for cell in ws[1]:
            cell.font = bold_font
        if column_widths:
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
        # Write data
        data = data_dict.get(sheet_name, [])  # Handle missing sheets gracefully
        if data:
            for col_num, (col_name, col_values) in enumerate(data.items(), start=1):
                for row, value in enumerate(col_values, start=2):
                    print(row, col_num, value)
                    ws.cell(row=row, column=col_num, value=value)

    del wb['Sheet']  # Remove default sheet
    with NamedTemporaryFile() as tmp:
        tmp.close() # with statement opened tmp, close it so wb.save can open it
        wb.save(tmp.name)
        with open(tmp.name, 'rb') as f:
            new_file_object = f.read()

    return FileResponse(BytesIO(new_file_object), as_attachment=True, filename=filename)